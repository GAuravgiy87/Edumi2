#!/usr/bin/env python
"""
Comprehensive Test Report Generator for EduMi 2.0
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def main():
    # 1. Define all features and test cases
    test_cases = [
        # ==================== ACCOUNTS ====================
        {
            "module": "Accounts",
            "submodule": "Authentication",
            "feature": "User Login",
            "description": "Verify user can log in with valid credentials",
            "test_steps": [
                "Navigate to login page",
                "Enter valid username",
                "Enter valid password",
                "Click Login button"
            ],
            "expected_result": "User is authenticated and redirected to home page",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "Authentication",
            "feature": "Invalid Login",
            "description": "Verify user cannot log in with invalid credentials",
            "test_steps": [
                "Navigate to login page",
                "Enter invalid username",
                "Enter invalid password",
                "Click Login button"
            ],
            "expected_result": "Error message is displayed and user remains on login page",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "Authentication",
            "feature": "User Registration",
            "description": "Verify new user can register an account",
            "test_steps": [
                "Navigate to register page",
                "Fill out registration form with valid data",
                "Click Register button"
            ],
            "expected_result": "New user account is created and user is redirected to home page",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "Authentication",
            "feature": "User Logout",
            "description": "Verify user can log out successfully",
            "test_steps": [
                "Ensure user is logged in",
                "Click Logout button/link"
            ],
            "expected_result": "User is logged out and redirected to login page",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # User Profiles
        {
            "module": "Accounts",
            "submodule": "User Profiles",
            "feature": "View Own Profile",
            "description": "Verify user can view their own profile",
            "test_steps": [
                "Log in as any user",
                "Navigate to profile page"
            ],
            "expected_result": "User's profile is displayed with all relevant information",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "User Profiles",
            "feature": "Edit Own Profile",
            "description": "Verify user can edit their own profile",
            "test_steps": [
                "Log in as any user",
                "Navigate to edit profile page",
                "Update profile fields (bio, display name, etc.)",
                "Save changes"
            ],
            "expected_result": "Profile information is updated successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Student Dashboard
        {
            "module": "Accounts",
            "submodule": "Student Dashboard",
            "feature": "Access Student Dashboard",
            "description": "Verify student can access their dashboard",
            "test_steps": [
                "Log in as a student",
                "Navigate to student dashboard"
            ],
            "expected_result": "Student dashboard is loaded and displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Teacher Dashboard
        {
            "module": "Accounts",
            "submodule": "Teacher Dashboard",
            "feature": "Access Teacher Dashboard",
            "description": "Verify teacher can access their dashboard",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to teacher dashboard"
            ],
            "expected_result": "Teacher dashboard is loaded and displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Admin Panel
        {
            "module": "Accounts",
            "submodule": "Admin Panel",
            "feature": "Access Admin Panel",
            "description": "Verify admin can access admin panel",
            "test_steps": [
                "Log in as an admin",
                "Navigate to admin panel"
            ],
            "expected_result": "Admin panel is loaded and displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "Admin Panel",
            "feature": "User Management",
            "description": "Verify admin can manage users (view, delete)",
            "test_steps": [
                "Log in as an admin",
                "Navigate to user management page",
                "View list of users",
                "Delete a test user (if applicable)"
            ],
            "expected_result": "User management operations are successful",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Messaging
        {
            "module": "Accounts",
            "submodule": "Messaging",
            "feature": "View Inbox",
            "description": "Verify user can view their inbox",
            "test_steps": [
                "Log in as any user",
                "Navigate to inbox"
            ],
            "expected_result": "Inbox is loaded and displays conversations",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "Messaging",
            "feature": "Start New Conversation",
            "description": "Verify user can start a new conversation",
            "test_steps": [
                "Log in as any user",
                "Navigate to user directory",
                "Select another user",
                "Start conversation"
            ],
            "expected_result": "New conversation is created",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "Messaging",
            "feature": "Send Message",
            "description": "Verify user can send messages in a conversation",
            "test_steps": [
                "Log in as any user",
                "Navigate to an existing conversation",
                "Type a message",
                "Send message"
            ],
            "expected_result": "Message is sent and displayed in conversation",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Notifications
        {
            "module": "Accounts",
            "submodule": "Notifications",
            "feature": "View Notifications",
            "description": "Verify user can view notifications",
            "test_steps": [
                "Log in as any user",
                "Navigate to notifications page"
            ],
            "expected_result": "Notifications are displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Accounts",
            "submodule": "Notifications",
            "feature": "Mark Notification as Read",
            "description": "Verify user can mark a notification as read",
            "test_steps": [
                "Log in as any user",
                "Navigate to notifications page",
                "Click on a notification to mark as read"
            ],
            "expected_result": "Notification is marked as read",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== MEETINGS ====================
        {
            "module": "Meetings",
            "submodule": "Classrooms",
            "feature": "Create Classroom",
            "description": "Verify teacher can create a new classroom",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to classroom creation page",
                "Fill out classroom details (title, description, password)",
                "Submit form"
            ],
            "expected_result": "New classroom is created successfully",
            "status": "Working",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Classrooms",
            "feature": "View Own Classrooms (Teacher)",
            "description": "Verify teacher can view their classrooms",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to teacher classrooms page"
            ],
            "expected_result": "Teacher's classrooms are listed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Classrooms",
            "feature": "View Own Classrooms (Student)",
            "description": "Verify student can view their classrooms",
            "test_steps": [
                "Log in as a student",
                "Navigate to student classrooms page"
            ],
            "expected_result": "Student's classrooms are listed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Classrooms",
            "feature": "Request to Join Classroom",
            "description": "Verify student can request to join a classroom",
            "test_steps": [
                "Log in as a student",
                "Find a classroom to join",
                "Submit join request"
            ],
            "expected_result": "Join request is sent to classroom teacher",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Classrooms",
            "feature": "Approve Join Request",
            "description": "Verify teacher can approve join requests",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to a classroom's detail page",
                "View pending join requests",
                "Approve a request"
            ],
            "expected_result": "Join request is approved and student is added to classroom",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Classrooms",
            "feature": "Deny Join Request",
            "description": "Verify teacher can deny join requests",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to a classroom's detail page",
                "View pending join requests",
                "Deny a request"
            ],
            "expected_result": "Join request is denied",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Meetings (LiveKit)
        {
            "module": "Meetings",
            "submodule": "Meetings (LiveKit)",
            "feature": "Create Meeting",
            "description": "Verify teacher can create a new meeting",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to create meeting page",
                "Fill out meeting details (title, time, duration)",
                "Submit form"
            ],
            "expected_result": "New meeting is created successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Meetings (LiveKit)",
            "feature": "Start Classroom Meeting",
            "description": "Verify teacher can start a meeting from a classroom",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to a classroom's detail page",
                "Click Start Meeting button"
            ],
            "expected_result": "Meeting is started and teacher is redirected to meeting room",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Meetings (LiveKit)",
            "feature": "Join Meeting",
            "description": "Verify user can join a live meeting",
            "test_steps": [
                "Obtain meeting code or navigate to join link",
                "Enter meeting code (if needed)",
                "Click Join Meeting button"
            ],
            "expected_result": "User joins the meeting room successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Meetings (LiveKit)",
            "feature": "End Meeting",
            "description": "Verify teacher can end a meeting",
            "test_steps": [
                "Be in a live meeting as teacher",
                "Click End Meeting button"
            ],
            "expected_result": "Meeting is ended for all participants",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Meeting Controls
        {
            "module": "Meetings",
            "submodule": "Meeting Controls",
            "feature": "Global Mute",
            "description": "Verify teacher can mute all participants",
            "test_steps": [
                "Be in a live meeting as teacher",
                "Click Global Mute button"
            ],
            "expected_result": "All participants are muted",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Meeting Controls",
            "feature": "Global Camera Off",
            "description": "Verify teacher can turn off all participants' cameras",
            "test_steps": [
                "Be in a live meeting as teacher",
                "Click Global Camera Off button"
            ],
            "expected_result": "All participants' cameras are turned off",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Meeting Controls",
            "feature": "Kick Participant",
            "description": "Verify teacher can kick a participant",
            "test_steps": [
                "Be in a live meeting as teacher",
                "Select a participant",
                "Click Kick button"
            ],
            "expected_result": "Participant is kicked from the meeting",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Meetings",
            "submodule": "Meeting Controls",
            "feature": "Meeting Chat",
            "description": "Verify participants can send and receive chat messages",
            "test_steps": [
                "Be in a live meeting",
                "Type a message in chat",
                "Send message"
            ],
            "expected_result": "Chat message is sent and visible to all participants",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== ASSIGNMENTS ====================
        {
            "module": "Assignments",
            "submodule": "Classroom Assignments",
            "feature": "View Classrooms with Quizzes Button",
            "description": "Verify classroom detail page has Quizzes button alongside Assignments",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to classroom detail page"
            ],
            "expected_result": "Quizzes button is visible next to Assignments button",
            "status": "Working",
            "notes": ""
        },
        {
            "module": "Assignments",
            "submodule": "Assignments Management",
            "feature": "Create Assignment",
            "description": "Verify teacher can create a new assignment with files/folders/links",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to classroom assignments page",
                "Click Create Assignment",
                "Fill out assignment details (title, description, due date)",
                "Add file attachments/folder uploads/link attachments",
                "Submit form"
            ],
            "expected_result": "Assignment is created successfully with all attachments",
            "status": "Working",
            "notes": "AssignmentQuestionFile model updated to support link_url and file_type fields"
        },
        {
            "module": "Assignments",
            "submodule": "Assignments Management",
            "feature": "View Assignments List",
            "description": "Verify teacher/student can view list of assignments for a classroom",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to classroom assignments page"
            ],
            "expected_result": "All assignments are listed with details",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Assignments",
            "submodule": "Assignments Management",
            "feature": "Submit Assignment",
            "description": "Verify student can submit an assignment with files/folders/links",
            "test_steps": [
                "Log in as a student",
                "Navigate to an assignment's detail page",
                "Add file/folder/link attachments",
                "Submit assignment"
            ],
            "expected_result": "Assignment submission is saved successfully with all attachments",
            "status": "Working",
            "notes": "AssignmentSubmissionFile model updated to support link_url and file_type fields"
        },
        {
            "module": "Assignments",
            "submodule": "Assignments Management",
            "feature": "View Assignment Details",
            "description": "Verify teacher/student can view full assignment details",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to an assignment's detail page"
            ],
            "expected_result": "Full assignment details and attachments are displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== QUIZZES ====================
        {
            "module": "Quizzes",
            "submodule": "Quizzes Management",
            "feature": "View Classroom Quizzes List",
            "description": "Verify teacher/student can view list of quizzes for a classroom (students see only published)",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to classroom quizzes page"
            ],
            "expected_result": "Quizzes are listed; students see only published, teachers see all",
            "status": "Working",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quizzes Management",
            "feature": "Create Quiz",
            "description": "Verify teacher can create a new quiz (saved as draft)",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to classroom quizzes page",
                "Click Create Quiz",
                "Fill out quiz details (title, description, due date, total marks)",
                "Save as draft"
            ],
            "expected_result": "Quiz is created successfully in draft status",
            "status": "Working",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quizzes Management",
            "feature": "Edit Quiz & Add Questions",
            "description": "Verify teacher can edit quiz details and add/delete questions",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to edit quiz page",
                "Update quiz details",
                "Click Add Question",
                "Fill out question details (type: MCQ/Text, question text, marks)",
                "Add choices for MCQ with correct answer",
                "Save question"
            ],
            "expected_result": "Quiz details are updated; question is added/removed successfully",
            "status": "Working",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quizzes Management",
            "feature": "Save Quiz as Draft",
            "description": "Verify teacher can save changes to quiz without publishing",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to edit quiz page",
                "Make changes",
                "Click Save Changes"
            ],
            "expected_result": "Changes are saved; quiz remains in current status; stays on edit page",
            "status": "Working",
            "notes": "edit_quiz view updated to redirect back to edit page on save"
        },
        {
            "module": "Quizzes",
            "submodule": "Quizzes Management",
            "feature": "Publish Quiz",
            "description": "Verify teacher can publish a draft quiz",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to edit quiz page (draft quiz)",
                "Click Publish Quiz"
            ],
            "expected_result": "Quiz status changes to published; students can now see/take it",
            "status": "Working",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quizzes Management",
            "feature": "Archive Quiz",
            "description": "Verify teacher can archive a published quiz",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to edit quiz page (published quiz)",
                "Click Archive Quiz"
            ],
            "expected_result": "Quiz status changes to archived",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quiz Taking",
            "feature": "View Quiz Details (Student)",
            "description": "Verify student can view published quiz details and questions before taking",
            "test_steps": [
                "Log in as a student",
                "Navigate to a published quiz's detail page"
            ],
            "expected_result": "Quiz details and questions are visible (correct answers hidden)",
            "status": "Working",
            "notes": "Fixed: added is_student to quiz_detail view, added missing endif in template"
        },
        {
            "module": "Quizzes",
            "submodule": "Quiz Taking",
            "feature": "Take Quiz",
            "description": "Verify student can take and submit a quiz",
            "test_steps": [
                "Log in as a student",
                "Navigate to a published quiz's detail page",
                "Click Take Quiz",
                "Answer all questions",
                "Submit quiz"
            ],
            "expected_result": "Quiz submission is saved successfully",
            "status": "Working",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quiz Evaluation",
            "feature": "View Quiz Submissions (Teacher)",
            "description": "Verify teacher can view all student submissions for a quiz",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to a quiz's detail page"
            ],
            "expected_result": "All student submissions are listed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quiz Evaluation",
            "feature": "Evaluate Quiz Submission",
            "description": "Verify teacher can evaluate (grade) a student's quiz submission",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to a student's quiz submission page",
                "Assign marks per question",
                "Add feedback",
                "Save evaluation"
            ],
            "expected_result": "Quiz evaluation is saved, student can see results",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Quizzes",
            "submodule": "Quiz Taking",
            "feature": "View Quiz Results (Student)",
            "description": "Verify student can view their quiz results after evaluation",
            "test_steps": [
                "Log in as a student",
                "Navigate to a quiz's detail page (submitted and evaluated)"
            ],
            "expected_result": "Student sees their marks, feedback, and answers",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== ATTENDANCE ====================
        {
            "module": "Attendance",
            "submodule": "Face Recognition",
            "feature": "Face Registration",
            "description": "Verify student can register their face",
            "test_steps": [
                "Log in as a student",
                "Navigate to face setup page",
                "Follow face registration instructions",
                "Complete registration"
            ],
            "expected_result": "Face is registered successfully for attendance",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Attendance",
            "submodule": "Face Recognition",
            "feature": "Face Detection During Meeting",
            "description": "Verify face recognition works during meetings",
            "test_steps": [
                "Join a meeting as a registered student",
                "Ensure camera is active",
                "Wait for attendance to be marked"
            ],
            "expected_result": "Attendance is marked as present when face is detected",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Attendance Reports
        {
            "module": "Attendance",
            "submodule": "Attendance Reports",
            "feature": "View Attendance Report",
            "description": "Verify teacher can view attendance reports",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to attendance reports page",
                "Select a classroom/meeting"
            ],
            "expected_result": "Attendance report is displayed with student attendance statuses",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Attendance",
            "submodule": "Attendance Reports",
            "feature": "Override Attendance",
            "description": "Verify teacher can override attendance records",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to attendance reports page",
                "Find an attendance record",
                "Change attendance status manually"
            ],
            "expected_result": "Attendance record is updated successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Attendance",
            "submodule": "Engagement Analytics",
            "feature": "View Engagement Report",
            "description": "Verify engagement report is available after meeting",
            "test_steps": [
                "Log in as a teacher",
                "Wait for a meeting to end",
                "Navigate to engagement report page"
            ],
            "expected_result": "Engagement report is generated and displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== CAMERAS ====================
        {
            "module": "Cameras",
            "submodule": "RTSP Cameras",
            "feature": "Add RTSP Camera",
            "description": "Verify admin can add an RTSP camera",
            "test_steps": [
                "Log in as an admin",
                "Navigate to add camera page",
                "Enter camera details (name, IP, port, etc.)",
                "Submit form"
            ],
            "expected_result": "RTSP camera is added successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "RTSP Cameras",
            "feature": "View Camera Feed",
            "description": "Verify camera feed is accessible",
            "test_steps": [
                "Log in as admin/authorized teacher",
                "Navigate to a camera's detail page",
                "View live feed"
            ],
            "expected_result": "Live camera feed is displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "Camera Permissions",
            "feature": "Grant Camera Permission",
            "description": "Verify admin can grant camera access to teachers",
            "test_steps": [
                "Log in as an admin",
                "Navigate to camera permissions page",
                "Grant permission to a teacher"
            ],
            "expected_result": "Teacher is granted access to the camera",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Camera Streaming & Recording
        {
            "module": "Cameras",
            "submodule": "Streaming & Recording",
            "feature": "Start Camera Streaming",
            "description": "Verify teacher can start camera streaming",
            "test_steps": [
                "Log in as an authorized teacher",
                "Navigate to camera dashboard",
                "Click Start Streaming button"
            ],
            "expected_result": "Camera streaming is started",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "Streaming & Recording",
            "feature": "Start Camera Recording",
            "description": "Verify teacher can start camera recording",
            "test_steps": [
                "Log in as an authorized teacher",
                "Navigate to camera dashboard",
                "Click Start Recording button"
            ],
            "expected_result": "Camera recording is started",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "Streaming & Recording",
            "feature": "Stop Camera Recording",
            "description": "Verify teacher can stop camera recording",
            "test_steps": [
                "Log in as an authorized teacher",
                "Navigate to camera dashboard",
                "Click Stop Recording button"
            ],
            "expected_result": "Camera recording is stopped and saved",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # Recordings
        {
            "module": "Cameras",
            "submodule": "Recordings",
            "feature": "View Recordings",
            "description": "Verify user can view saved recordings",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to recordings page"
            ],
            "expected_result": "Recordings are listed and accessible",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "Recordings",
            "feature": "Watch Recording",
            "description": "Verify user can watch a recording",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to recordings page",
                "Click on a recording to watch"
            ],
            "expected_result": "Recording video is played successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "Recordings",
            "feature": "Publish/Unpublish Recording",
            "description": "Verify teacher can publish/unpublish recordings",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to recordings page",
                "Toggle publish status for a recording"
            ],
            "expected_result": "Recording publish status is updated",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "Recordings",
            "feature": "Delete Recording",
            "description": "Verify teacher can delete recordings",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to recordings page",
                "Click Delete button for a recording"
            ],
            "expected_result": "Recording is deleted successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== MOBILE CAMERAS ====================
        {
            "module": "Mobile Cameras",
            "submodule": "Mobile Cameras",
            "feature": "Add Mobile Camera",
            "description": "Verify admin can add a mobile camera (IP Webcam/DroidCam)",
            "test_steps": [
                "Log in as an admin",
                "Navigate to add mobile camera page",
                "Enter camera details",
                "Submit form"
            ],
            "expected_result": "Mobile camera is added successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Mobile Cameras",
            "submodule": "Mobile Cameras",
            "feature": "View Mobile Camera Feed",
            "description": "Verify mobile camera feed is accessible",
            "test_steps": [
                "Log in as admin/authorized teacher",
                "Navigate to a mobile camera's detail page",
                "View live feed"
            ],
            "expected_result": "Live mobile camera feed is displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== HEAD COUNT ====================
        {
            "module": "Head Count",
            "submodule": "Head Count",
            "feature": "Start Head Count Session",
            "description": "Verify head count session can be started",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to head count dashboard",
                "Click Start Head Count button"
            ],
            "expected_result": "Head count session is started",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Head Count",
            "submodule": "Head Count",
            "feature": "View Head Count Logs",
            "description": "Verify head count logs are viewable",
            "test_steps": [
                "Log in as a teacher/admin",
                "Navigate to head count logs page"
            ],
            "expected_result": "Head count logs are displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Head Count",
            "submodule": "Head Count",
            "feature": "Export Head Count CSV",
            "description": "Verify head count logs can be exported as CSV",
            "test_steps": [
                "Log in as a teacher/admin",
                "Navigate to head count logs page",
                "Click Export CSV button"
            ],
            "expected_result": "CSV file with head count logs is downloaded",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== VIDEOS ====================
        {
            "module": "Videos",
            "submodule": "Video Management",
            "feature": "Upload Video",
            "description": "Verify user can upload a video",
            "test_steps": [
                "Log in as a teacher/admin",
                "Navigate to video upload page",
                "Select a video file",
                "Fill out video details",
                "Upload video"
            ],
            "expected_result": "Video is uploaded successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Videos",
            "submodule": "Video Management",
            "feature": "View Video List",
            "description": "Verify user can view list of videos",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to video list page"
            ],
            "expected_result": "List of videos is displayed",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Videos",
            "submodule": "Video Management",
            "feature": "Watch Video",
            "description": "Verify user can watch a video",
            "test_steps": [
                "Log in as a teacher/student",
                "Navigate to a video's detail page",
                "Play video"
            ],
            "expected_result": "Video is played successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Videos",
            "submodule": "Video Management",
            "feature": "Edit Video Details",
            "description": "Verify user can edit video details",
            "test_steps": [
                "Log in as a teacher/admin",
                "Navigate to a video's edit page",
                "Update video details (title, description)",
                "Save changes"
            ],
            "expected_result": "Video details are updated successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Videos",
            "submodule": "Video Management",
            "feature": "Delete Video",
            "description": "Verify user can delete a video",
            "test_steps": [
                "Log in as a teacher/admin",
                "Navigate to video list page",
                "Click Delete button for a video"
            ],
            "expected_result": "Video is deleted successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== VIDEO EDITING ====================
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Create Edit Session",
            "description": "Verify user can create a video edit session",
            "test_steps": [
                "Log in as a teacher/admin",
                "Navigate to a video's edit page",
                "Start new edit session"
            ],
            "expected_result": "Edit session is created",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Trim Video",
            "description": "Verify user can trim a video",
            "test_steps": [
                "In a video edit session",
                "Select trim action",
                "Set start and end times",
                "Apply action"
            ],
            "expected_result": "Trim action is added to edit session",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Mute Video Audio",
            "description": "Verify user can mute video audio",
            "test_steps": [
                "In a video edit session",
                "Select mute action",
                "Apply action"
            ],
            "expected_result": "Mute action is added to edit session",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Rotate Video",
            "description": "Verify user can rotate a video",
            "test_steps": [
                "In a video edit session",
                "Select rotate action",
                "Set rotation angle",
                "Apply action"
            ],
            "expected_result": "Rotate action is added to edit session",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Add Text Overlay",
            "description": "Verify user can add text overlay to video",
            "test_steps": [
                "In a video edit session",
                "Select add text action",
                "Enter text and set position",
                "Apply action"
            ],
            "expected_result": "Text overlay action is added to edit session",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Add Audio Overlay",
            "description": "Verify user can add audio overlay to video",
            "test_steps": [
                "In a video edit session",
                "Select add audio action",
                "Upload audio file",
                "Apply action"
            ],
            "expected_result": "Add audio action is added to edit session",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Process Edits",
            "description": "Verify user can process edits and generate final video",
            "test_steps": [
                "In a video edit session with actions added",
                "Click Process button",
                "Wait for processing to complete"
            ],
            "expected_result": "Final edited video is generated successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Video Editing",
            "submodule": "Video Editing",
            "feature": "Download Edited Video",
            "description": "Verify user can download the final edited video",
            "test_steps": [
                "After edit processing completes",
                "Click Download button"
            ],
            "expected_result": "Edited video is downloaded",
            "status": "To Be Tested",
            "notes": ""
        },
        
        # ==================== ADDITIONAL ====================
        {
            "module": "Cameras",
            "submodule": "Recordings",
            "feature": "Edit Recording (Trim)",
            "description": "Verify user can trim a camera recording",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to edit recording page",
                "Set trim start/end times",
                "Apply trim"
            ],
            "expected_result": "Recording is trimmed successfully",
            "status": "To Be Tested",
            "notes": ""
        },
        {
            "module": "Cameras",
            "submodule": "Recordings",
            "feature": "Generate Recording Thumbnail",
            "description": "Verify thumbnail can be generated for a recording",
            "test_steps": [
                "Log in as a teacher",
                "Navigate to recording page",
                "Click Generate Thumbnail button"
            ],
            "expected_result": "Thumbnail is generated and displayed",
            "status": "To Be Tested",
            "notes": ""
        }
    ]
    
    # 2. Generate Markdown report
    generate_markdown_report(test_cases)
    
    # 3. Generate Excel report
    generate_excel_report(test_cases)
    
    print("Reports generated successfully!")
    print("- comprehensive_test_report.md")
    print("- comprehensive_test_report.xlsx")


def generate_markdown_report(test_cases):
    """Generate the comprehensive test report in Markdown format."""
    md_content = """# EduMi 2.0 - Comprehensive Test Report

## Overview
This report contains a detailed test plan for all features and modules of the EduMi 2.0 educational platform.

---

## Test Cases

"""
    
    # Group test cases by module
    from collections import defaultdict
    grouped_tests = defaultdict(list)
    for test in test_cases:
        grouped_tests[test["module"]].append(test)
    
    # Generate content for each module
    for module, tests in sorted(grouped_tests.items()):
        md_content += f"## {module}\n\n"
        
        # Group further by submodule
        sub_grouped = defaultdict(list)
        for test in tests:
            sub_grouped[test["submodule"]].append(test)
        
        for submodule, sub_tests in sorted(sub_grouped.items()):
            md_content += f"### {submodule}\n\n"
            
            # Table for this submodule
            md_content += "| Feature | Description | Status | Notes |\n"
            md_content += "|---------|-------------|--------|-------|\n"
            
            for test in sub_tests:
                status_emoji = {
                    "Passed": "✅",
                    "Failed": "❌",
                    "To Be Tested": "⏳",
                    "In Progress": "🔄"
                }.get(test["status"], "")
                md_content += f"| **{test['feature']}** | {test['description']} | {status_emoji} {test['status']} | {test['notes']} |\n"
            
            md_content += "\n"
            
            # Detailed test steps for each test case
            for idx, test in enumerate(sub_tests, 1):
                md_content += f"#### {idx}. {test['feature']}\n\n"
                md_content += f"**Description**: {test['description']}\n\n"
                md_content += "**Test Steps**:\n"
                for step in test["test_steps"]:
                    md_content += f"- {step}\n"
                md_content += f"\n**Expected Result**: {test['expected_result']}\n\n"
                md_content += f"**Status**: {test['status']}\n\n"
                if test["notes"]:
                    md_content += f"**Notes**: {test['notes']}\n\n"
                md_content += "---\n\n"
    
    # Write to file
    with open("comprehensive_test_report.md", "w", encoding="utf-8") as f:
        f.write(md_content)


def generate_excel_report(test_cases):
    """Generate the comprehensive test report in Excel format using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Headers
    headers = [
        "Module", "Submodule", "Feature", "Description", 
        "Test Steps", "Expected Result", "Status", "Notes"
    ]
    
    # Style for header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write test cases
    for row_idx, test in enumerate(test_cases, 2):
        ws.cell(row=row_idx, column=1, value=test["module"])
        ws.cell(row=row_idx, column=2, value=test["submodule"])
        ws.cell(row=row_idx, column=3, value=test["feature"])
        ws.cell(row=row_idx, column=4, value=test["description"])
        ws.cell(row=row_idx, column=5, value="\n".join(test["test_steps"]))
        ws.cell(row=row_idx, column=6, value=test["expected_result"])
        ws.cell(row=row_idx, column=7, value=test["status"])
        ws.cell(row=row_idx, column=8, value=test["notes"])
        
        # Wrap text for readability
        for col_idx in range(1, 9):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-fit columns (approximate)
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add filters
    ws.auto_filter.ref = ws.dimensions
    
    # Save the workbook
    wb.save("comprehensive_test_report.xlsx")


if __name__ == "__main__":
    main()
