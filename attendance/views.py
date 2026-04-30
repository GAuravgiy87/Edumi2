"""
Attendance module views — all endpoints for:
  • Student face registration (upload / camera capture)
  • Teacher controls (schedule, settings, override)
  • Attendance records & reports (daily, student, classroom)
  • Export (Excel)
"""
import json
import base64
import logging

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.utils import timezone
from django.contrib.auth.models import User
from django.db import transaction
from django.core.files.base import ContentFile

from meetings.models import Classroom, Meeting
from .models import (
    StudentFaceProfile, ClassSchedule, AttendanceRecord,
    FaceRecognitionLog, AttendanceSettings
)
from .forms import FacePhotoForm
from .face_service import FaceService, get_face_service
from .services import get_daily_report_context, get_classroom_attendance_stats

logger = logging.getLogger('attendance')


# ══════════════════════════════════════════════════════════════
#  STUDENT: FACE REGISTRATION
# ══════════════════════════════════════════════════════════════

@login_required
def face_setup(request):
    """Landing page for face registration — tabs: upload / camera capture."""
    profile = getattr(request.user, 'face_profile', None)
    u_profile = request.user.userprofile
    
    # Check if profile info is complete
    info_complete = all([u_profile.roll_number, u_profile.branch, u_profile.contact_number])
    
    ctx = {
        'has_profile': profile is not None and profile.is_active,
        'profile':     profile,
        'u_profile':   u_profile,
        'info_complete': info_complete,
        'page_title':  'Face Registration',
    }
    return render(request, 'attendance/face_setup.html', ctx)


@login_required
@require_POST
def upload_face_photo(request):
    """Handle file upload approach to face registration."""
    form = FacePhotoForm(request.POST, request.FILES)
    # Note: form might not be valid if fields are missing, but we'll check profile
    
    photo = request.FILES.get('photo')
    if not photo:
        messages.error(request, 'Please select a photo.')
        return redirect('face_setup')

    image_bytes = photo.read()
    svc = get_face_service()
    result = svc.extract_embedding(image_bytes)

    if result['status'] != 'success':
        messages.error(request, f"Face detection failed: {result['message']}")
        return redirect('face_setup')

    # Ensure profile info is present
    u_profile = request.user.userprofile
    roll = request.POST.get('roll_number') or u_profile.roll_number
    branch = request.POST.get('branch') or u_profile.branch
    contact = request.POST.get('contact_number') or u_profile.contact_number

    if not all([roll, branch, contact]):
        messages.error(request, 'Student details (Roll, Branch, Contact) are missing. Please complete your profile.')
        return redirect('face_setup')

    # Update profile if provided in request
    u_profile.roll_number = roll
    u_profile.branch = branch
    u_profile.contact_number = contact
    u_profile.save()

    encrypted, checksum = svc.prepare_for_storage(result['embedding'])
    photo_file = ContentFile(image_bytes, name=f"{request.user.username}_face.jpg")

    StudentFaceProfile.objects.update_or_create(
        student=request.user,
        defaults={
            'face_embedding_encrypted': encrypted,
            'embedding_checksum':       checksum,
            'face_quality_score':       result['quality'],
            'is_active':                True,
            'registration_ip':          _get_client_ip(request),
            'face_photo':               photo_file,
        }
    )

    messages.success(request, "✅ Face registered successfully!")
    return redirect('face_setup')


@login_required
@require_POST
def capture_face_photo(request):
    """Handle camera-captured base64 frame approach to face registration."""
    try:
        body   = json.loads(request.body)
        b64    = body.get('frame_b64', '')
        
        # Check profile first, then request
        u_profile = request.user.userprofile
        roll    = body.get('roll_number') or u_profile.roll_number
        branch  = body.get('branch') or u_profile.branch
        contact = body.get('contact_number') or u_profile.contact_number

        if not b64:
            return JsonResponse({'status': 'error', 'message': 'No frame data received.'}, status=400)
        
        if not all([roll, branch, contact]):
            return JsonResponse({'status': 'error', 'message': 'Student details are missing. Please complete your profile first.'}, status=400)

        if ',' in b64:
            b64 = b64.split(',', 1)[1]
        image_bytes = base64.b64decode(b64)
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Invalid request data: {exc}'}, status=400)

    svc    = get_face_service()
    result = svc.extract_embedding(image_bytes)

    if result['status'] != 'success':
        return JsonResponse({'status': 'error', 'message': result['message']})

    # Ensure profile is updated if data came in request
    u_profile.roll_number = roll
    u_profile.branch = branch
    u_profile.contact_number = contact
    u_profile.save()

    encrypted, checksum = svc.prepare_for_storage(result['embedding'])
    photo_file = ContentFile(image_bytes, name=f"{request.user.username}_face.jpg")

    StudentFaceProfile.objects.update_or_create(
        student=request.user,
        defaults={
            'face_embedding_encrypted': encrypted,
            'embedding_checksum':       checksum,
            'face_quality_score':       result['quality'],
            'is_active':                True,
            'registration_ip':          _get_client_ip(request),
            'face_photo':               photo_file,
        }
    )

    return JsonResponse({
        'status':  'success',
        'quality': result['quality'],
        'message': '✅ Face registered successfully!',
    })


@login_required
@require_POST
def detect_face(request):
    """
    Lightweight face detection for real-time feedback.
    Includes low-light enhancement.
    """
    try:
        body = json.loads(request.body)
        b64 = body.get('frame_b64', '')
        if not b64:
            return JsonResponse({'status': 'no_frame'})
        if ',' in b64:
            b64 = b64.split(',', 1)[1]
        image_bytes = base64.b64decode(b64)
    except Exception:
        return JsonResponse({'status': 'invalid_data'})

    import cv2
    import numpy as np
    from PIL import Image

    # Decode for enhancement
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    
    if img is None:
        return JsonResponse({'status': 'decode_error'})

    # ── Low-Light Enhancement ──
    # Check average brightness
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    avg_brightness = np.mean(gray)
    
    enhanced = False
    if avg_brightness < 60: # Threshold for "dark"
        # Apply CLAHE (Contrast Limited Adaptive Histogram Equalization)
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
        img_yuv = cv2.cvtColor(img, cv2.COLOR_BGR2YUV)
        img_yuv[:,:,0] = clahe.apply(img_yuv[:,:,0])
        img = cv2.cvtColor(img_yuv, cv2.COLOR_YUV2BGR)
        enhanced = True
        
        # Encode back to bytes for face_recognition
        _, buffer = cv2.imencode('.jpg', img)
        image_bytes = buffer.tobytes()

    svc = get_face_service()
    # Use hog model (fast) and skip liveness for detection feedback
    result = svc.extract_embedding(image_bytes, live=False)

    return JsonResponse({
        'status': result['status'],
        'quality': result['quality'],
        'low_light_enhanced': enhanced,
        'message': result['message']
    })


@login_required
@require_POST
def update_profile_info(request):
    """AJAX: Update student profile details (roll, branch, contact)."""
    try:
        body = json.loads(request.body)
        roll = body.get('roll_number', '')
        branch = body.get('branch', '')
        contact = body.get('contact_number', '')
        
        if not all([roll, branch, contact]):
            return JsonResponse({'status': 'error', 'message': 'All fields are required.'}, status=400)
            
        profile = request.user.userprofile
        profile.roll_number = roll
        profile.branch = branch
        profile.contact_number = contact
        profile.save()
        
        return JsonResponse({'status': 'success', 'message': 'Profile updated successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def face_registration_status(request):
    """AJAX: return whether the current user has a face profile."""
    profile = getattr(request.user, 'face_profile', None)
    return JsonResponse({
        'registered': profile is not None and profile.is_active,
        'quality':    profile.face_quality_score if profile else 0,
        'updated_at': profile.updated_at.isoformat() if profile else None,
    })


# ══════════════════════════════════════════════════════════════
#  STUDENT: VIEW OWN ATTENDANCE
# ══════════════════════════════════════════════════════════════

@login_required
def my_attendance(request):
    """Student view of their own attendance records across all classrooms."""
    records = AttendanceRecord.objects.filter(
        student=request.user
    ).select_related('meeting', 'classroom').order_by('-date')

    summary = {}
    for rec in records:
        key = rec.classroom_id
        if key not in summary:
            summary[key] = {
                'classroom': rec.classroom,
                'total': 0, 'present': 0, 'absent': 0, 'late': 0,
            }
        summary[key]['total'] += 1
        summary[key][rec.status] = summary[key].get(rec.status, 0) + 1

    for s in summary.values():
        s['percentage'] = round((s['present'] + s.get('late', 0)) / s['total'] * 100) if s['total'] else 0

    ctx = {
        'records':    records,
        'summary':    list(summary.values()),
        'page_title': 'My Attendance',
    }
    return render(request, 'attendance/my_attendance.html', ctx)


# ══════════════════════════════════════════════════════════════
#  TEACHER: CLASS SCHEDULE
# ══════════════════════════════════════════════════════════════

@login_required
def set_class_schedule(request, classroom_id):
    """Teacher sets which days of the week classes are held."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)

    if request.method == 'POST':
        ClassSchedule.objects.filter(classroom=classroom).delete()
        days        = request.POST.getlist('days')
        start_times = request.POST.getlist('start_times')
        end_times   = request.POST.getlist('end_times')

        for day, start, end in zip(days, start_times, end_times):
            if day and start and end:
                ClassSchedule.objects.create(
                    classroom=classroom,
                    day_of_week=int(day),
                    start_time=start,
                    end_time=end,
                    created_by=request.user,
                )
        messages.success(request, "Class schedule updated successfully.")
        return redirect('classroom_detail', classroom_id=classroom_id)

    schedules = ClassSchedule.objects.filter(classroom=classroom)
    ctx = {
        'classroom': classroom,
        'schedules': schedules,
        'all_days':  ClassSchedule.DAY_CHOICES,
        'page_title': f'Schedule — {classroom.title}',
    }
    return render(request, 'attendance/class_schedule.html', ctx)


# ══════════════════════════════════════════════════════════════
#  TEACHER: ATTENDANCE SETTINGS
# ══════════════════════════════════════════════════════════════

@login_required
def attendance_settings_view(request, classroom_id):
    """Teacher configures face-recognition thresholds for a classroom."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    settings_obj, _ = AttendanceSettings.objects.get_or_create(classroom=classroom)

    if request.method == 'POST':
        settings_obj.face_recognition_enabled    = 'face_recognition_enabled' in request.POST
        settings_obj.confidence_threshold         = float(request.POST.get('confidence_threshold', 0.55))
        settings_obj.presence_duration_seconds    = int(request.POST.get('presence_duration_seconds', 30))
        settings_obj.late_threshold_minutes       = int(request.POST.get('late_threshold_minutes', 10))
        settings_obj.recognition_interval_seconds = int(request.POST.get('recognition_interval_seconds', 15))
        settings_obj.enforce_schedule             = 'enforce_schedule' in request.POST
        settings_obj.save()
        messages.success(request, "Attendance settings saved.")
        return redirect('attendance_settings', classroom_id=classroom_id)

    ctx = {
        'classroom':    classroom,
        'att_settings': settings_obj,
        'page_title':   f'Attendance Settings — {classroom.title}',
    }
    return render(request, 'attendance/attendance_settings.html', ctx)


# ══════════════════════════════════════════════════════════════
#  TEACHER: OVERRIDE ATTENDANCE
# ══════════════════════════════════════════════════════════════

@login_required
@require_POST
def override_attendance(request, record_id):
    """Teacher manually marks a student present/absent/late."""
    record = get_object_or_404(AttendanceRecord, id=record_id)
    if record.classroom.teacher != request.user:
        return JsonResponse({'status': 'forbidden'}, status=403)

    new_status = request.POST.get('status', 'present')
    reason     = request.POST.get('reason', '')

    if new_status not in dict(AttendanceRecord.STATUS_CHOICES):
        return JsonResponse({'status': 'error', 'message': 'Invalid status.'}, status=400)

    record.status               = new_status
    record.override_reason      = reason
    record.overridden_by        = request.user
    record.verification_method  = 'manual'
    record.save()

    return JsonResponse({'status': 'success', 'new_status': new_status})


# ══════════════════════════════════════════════════════════════
#  TEACHER: REPORTS
# ══════════════════════════════════════════════════════════════

@login_required
def daily_report(request, classroom_id):
    """Show attendance for a specific date (default: today)."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    date_str  = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))

    try:
        from datetime import date as ddate
        report_date = ddate.fromisoformat(date_str)
    except ValueError:
        report_date = timezone.now().date()

    report_ctx = get_daily_report_context(classroom, report_date)
    ctx = {
        'classroom':     classroom,
        'report_date':   report_date,
        'date_str':      date_str,
        'page_title':    f'Daily Attendance — {report_date}',
        **report_ctx
    }
    return render(request, 'attendance/daily_report.html', ctx)


@login_required
def student_report(request, classroom_id, student_id):
    """Per-student attendance history within a classroom."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    student   = get_object_or_404(User, id=student_id)

    records = AttendanceRecord.objects.filter(
        classroom=classroom, student=student
    ).select_related('meeting').order_by('-date')

    total   = records.count()
    present = records.filter(status__in=['present', 'late']).count()
    pct     = round(present / total * 100) if total else 0

    ctx = {
        'classroom':  classroom,
        'student':    student,
        'records':    records,
        'total':      total,
        'present':    present,
        'absent':     total - present,
        'percentage': pct,
        'page_title': f'Attendance — {student.get_full_name() or student.username}',
    }
    return render(request, 'attendance/student_report.html', ctx)


@login_required
def classroom_attendance_overview(request, classroom_id):
    """Teacher dashboard: all students with overall attendance %."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    total_meetings = Meeting.objects.filter(classroom=classroom, status='ended').count()

    rows = get_classroom_attendance_stats(classroom)

    face_registered_count = sum(1 for r in rows if r['face_registered'])
    settings_obj, _ = AttendanceSettings.objects.get_or_create(classroom=classroom)
    schedules = ClassSchedule.objects.filter(classroom=classroom, is_active=True)

    ctx = {
        'classroom':             classroom,
        'rows':                  rows,
        'att_settings':          settings_obj,
        'schedules':             schedules,
        'total_meetings':        total_meetings,
        'face_registered_count': face_registered_count,
        'page_title':            f'Attendance — {classroom.title}',
    }
    return render(request, 'attendance/classroom_overview.html', ctx)


# ══════════════════════════════════════════════════════════════
#  EXPORT
# ══════════════════════════════════════════════════════════════

@login_required
def export_excel(request, classroom_id):
    """Export all attendance records for the classroom as Excel."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect('classroom_attendance_overview', classroom_id=classroom_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    header_fill  = PatternFill('solid', fgColor='1877F2')
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    status_colors = {
        'present': 'D4EDDA', 'late': 'FFF3CD',
        'absent':  'F8D7DA', 'partial': 'D1ECF1',
    }

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'Attendance Report  —  {classroom.title}'
    title_cell.font  = Font(bold=True, size=14, color='1A1A2E')
    title_cell.alignment = center_align

    headers = ['Student Name', 'Student ID', 'Date', 'Meeting', 'Status',
               'Time In', 'Method', 'Confidence']
    ws.append([])
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 14

    records = AttendanceRecord.objects.filter(
        classroom=classroom
    ).select_related('student', 'student__userprofile', 'meeting').order_by('-date', 'student__last_name')

    for row_idx, rec in enumerate(records, 4):
        try:
            sid = rec.student.userprofile.student_id or '—'
        except Exception:
            sid = '—'

        time_in = rec.marked_present_at.strftime('%H:%M:%S') if rec.marked_present_at else '—'
        conf    = f"{rec.face_match_confidence * 100:.1f}%" if rec.face_match_confidence else '—'

        row_data = [
            rec.student.get_full_name() or rec.student.username,
            sid,
            str(rec.date),
            rec.meeting.title,
            rec.get_status_display(),
            time_in,
            rec.get_verification_method_display(),
            conf,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if col_idx == 5:
                color = status_colors.get(rec.status, 'FFFFFF')
                cell.fill = PatternFill('solid', fgColor=color)
                cell.alignment = center_align
                cell.font = Font(bold=True)

    filename = f"attendance_{classroom.class_code}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ══════════════════════════════════════════════════════════════
#  API: CHECK SCHEDULE (used by JS before starting FR)
# ══════════════════════════════════════════════════════════════

@login_required
def check_schedule_api(request, meeting_code):
    """Return whether today is a scheduled class day for this meeting's classroom."""
    try:
        meeting   = Meeting.objects.select_related('classroom').get(meeting_code=meeting_code)
        classroom = meeting.classroom
    except Meeting.DoesNotExist:
        return JsonResponse({'scheduled': False, 'message': 'Meeting not found.'})

    settings_obj, _ = AttendanceSettings.objects.get_or_create(classroom=classroom)
    if not settings_obj.face_recognition_enabled:
        return JsonResponse({'scheduled': False, 'message': 'Face recognition disabled for this classroom.'})

    if not settings_obj.enforce_schedule:
        return JsonResponse({'scheduled': True, 'message': 'Schedule not enforced.'})

    today = timezone.localdate()
    scheduled = ClassSchedule.objects.filter(
        classroom=classroom,
        day_of_week=today.weekday(),
        is_active=True
    ).exists()

    return JsonResponse({
        'scheduled': scheduled,
        'message':   'Class is scheduled today.' if scheduled else 'No class scheduled today — attendance not recorded.',
        'interval':  settings_obj.recognition_interval_seconds,
    })


# ══════════════════════════════════════════════════════════════
#  TEACHER: ENGAGEMENT REPORT
# ══════════════════════════════════════════════════════════════

@login_required
def engagement_report_view(request, meeting_id):
    """Teacher views the engagement report for a completed meeting."""
    meeting = get_object_or_404(Meeting, id=meeting_id)
    if meeting.teacher != request.user and not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access denied.")

    from .models import EngagementReport
    report = EngagementReport.objects.filter(meeting=meeting).first()

    if not report and meeting.status == 'ended':
        from .engagement_service import generate_engagement_report
        generate_engagement_report(meeting.id)
        report = EngagementReport.objects.filter(meeting=meeting).first()

    import os
    from django.conf import settings
    log_filename = f'engagement_{meeting.meeting_code}.csv'
    log_path = os.path.join(settings.MEDIA_ROOT, 'meeting_logs', log_filename)
    log_exists = os.path.exists(log_path)

    ctx = {
        'meeting':    meeting,
        'report':     report,
        'page_title': f'Engagement Report — {meeting.title}',
        'log_exists': log_exists,
        'log_url':    f"{settings.MEDIA_URL}meeting_logs/{log_filename}" if log_exists else None,
    }
    return render(request, 'attendance/engagement_report.html', ctx)


# ══════════════════════════════════════════════════════════════
#  ADMIN: STUDENT FACE PHOTOS
# ══════════════════════════════════════════════════════════════

@login_required
def admin_face_photos(request):
    """Admin-only view of all student face registration photos."""
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access denied.")

    profiles = StudentFaceProfile.objects.filter(
        is_active=True, face_photo__isnull=False
    ).exclude(face_photo='').select_related('student', 'student__userprofile').order_by('-updated_at')

    ctx = {
        'profiles':   profiles,
        'page_title': 'Student Face Photos',
    }
    return render(request, 'attendance/admin_face_photos.html', ctx)


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════

def _get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')
