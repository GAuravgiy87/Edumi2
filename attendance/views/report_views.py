"""
Attendance report views + export + API + engagement:
daily_report, student_report, classroom_overview, export_excel,
check_schedule_api, engagement_report_view, admin_face_photos.
"""
import logging
import os
from datetime import date as ddate

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.views.decorators.http import require_GET
from django.utils import timezone
from django.contrib.auth.models import User
from django.conf import settings

from meetings.models import Classroom, Meeting
from attendance.models import (
    StudentFaceProfile, ClassSchedule, AttendanceRecord, AttendanceSettings,
)
from attendance.services import get_daily_report_context, get_classroom_attendance_stats

logger = logging.getLogger('attendance')


@login_required
def daily_report(request, classroom_id):
    """Show attendance for a specific date (default: today)."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    date_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
    try:
        report_date = ddate.fromisoformat(date_str)
    except ValueError:
        report_date = timezone.now().date()

    report_ctx = get_daily_report_context(classroom, report_date)
    return render(request, 'attendance/daily_report.html', {
        'classroom': classroom, 'report_date': report_date, 'date_str': date_str,
        'page_title': f'Daily Attendance — {report_date}', **report_ctx,
    })


@login_required
def student_report(request, classroom_id, student_id):
    """Per-student attendance history within a classroom."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    student = get_object_or_404(User, id=student_id)
    records = AttendanceRecord.objects.filter(
        classroom=classroom, student=student
    ).select_related('meeting').order_by('-date')
    total = records.count()
    present = records.filter(status__in=['present', 'late']).count()
    pct = round(present / total * 100) if total else 0
    return render(request, 'attendance/student_report.html', {
        'classroom': classroom, 'student': student, 'records': records,
        'total': total, 'present': present, 'absent': total - present,
        'percentage': pct, 'page_title': f'Attendance — {student.get_full_name() or student.username}',
    })


@login_required
def classroom_attendance_overview(request, classroom_id):
    """Teacher dashboard: all students with overall attendance percentage."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    total_meetings = Meeting.objects.filter(classroom=classroom, status='ended').count()
    rows = get_classroom_attendance_stats(classroom)
    face_registered_count = sum(1 for r in rows if r['face_registered'])
    settings_obj, _ = AttendanceSettings.objects.get_or_create(classroom=classroom)
    schedules = ClassSchedule.objects.filter(classroom=classroom, is_active=True)
    return render(request, 'attendance/classroom_overview.html', {
        'classroom': classroom, 'rows': rows, 'att_settings': settings_obj,
        'schedules': schedules, 'total_meetings': total_meetings,
        'face_registered_count': face_registered_count,
        'page_title': f'Attendance — {classroom.title}',
    })


@login_required
def export_excel(request, classroom_id):
    """Export all attendance records for the classroom as an Excel file."""
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user)
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect('classroom_attendance_overview', classroom_id=classroom_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    header_fill = PatternFill('solid', fgColor='1877F2')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    status_colors = {'present': 'D4EDDA', 'late': 'FFF3CD', 'absent': 'F8D7DA', 'partial': 'D1ECF1'}

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'Attendance Report  —  {classroom.title}'
    title_cell.font = Font(bold=True, size=14, color='1A1A2E')
    title_cell.alignment = center_align

    headers = ['Student Name', 'Student ID', 'Date', 'Meeting', 'Status', 'Time In', 'Method', 'Confidence']
    ws.append([])
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin

    for col, width in zip('ABCDEFGH', [24, 14, 14, 28, 12, 12, 20, 14]):
        ws.column_dimensions[col].width = width

    records = AttendanceRecord.objects.filter(classroom=classroom).select_related(
        'student', 'student__userprofile', 'meeting'
    ).order_by('-date', 'student__last_name')

    for row_idx, rec in enumerate(records, 4):
        try:
            sid = rec.student.userprofile.student_id or '—'
        except Exception:
            sid = '—'
        time_in = rec.marked_present_at.strftime('%H:%M:%S') if rec.marked_present_at else '—'
        conf = f"{rec.face_match_confidence * 100:.1f}%" if rec.face_match_confidence else '—'
        row_data = [
            rec.student.get_full_name() or rec.student.username, sid, str(rec.date),
            rec.meeting.title, rec.get_status_display(), time_in,
            rec.get_verification_method_display(), conf,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical='center')
            if col_idx == 5:
                cell.fill = PatternFill('solid', fgColor=status_colors.get(rec.status, 'FFFFFF'))
                cell.alignment = center_align
                cell.font = Font(bold=True)

    filename = f"attendance_{classroom.class_code}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def check_schedule_api(request, meeting_code):
    """Return whether today is a scheduled class day for this meeting's classroom."""
    try:
        meeting = Meeting.objects.select_related('classroom').get(meeting_code=meeting_code)
        classroom = meeting.classroom
    except Meeting.DoesNotExist:
        return JsonResponse({'scheduled': False, 'message': 'Meeting not found.'})

    settings_obj, _ = AttendanceSettings.objects.get_or_create(classroom=classroom)
    if not settings_obj.face_recognition_enabled:
        return JsonResponse({'scheduled': False, 'message': 'Face recognition disabled for this classroom.'})
    if not settings_obj.enforce_schedule:
        return JsonResponse({'scheduled': True, 'message': 'Schedule not enforced.'})

    today = timezone.localdate()
    scheduled = ClassSchedule.objects.filter(
        classroom=classroom, day_of_week=today.weekday(), is_active=True
    ).exists()
    return JsonResponse({
        'scheduled': scheduled,
        'message': 'Class is scheduled today.' if scheduled else 'No class scheduled today — attendance not recorded.',
        'interval': settings_obj.recognition_interval_seconds,
    })


@login_required
def engagement_report_view(request, meeting_id):
    """Teacher views the engagement report for a completed meeting."""
    meeting = get_object_or_404(Meeting, id=meeting_id)
    if meeting.teacher != request.user and not request.user.is_superuser:
        return HttpResponseForbidden("Access denied.")

    from attendance.models import EngagementReport
    report = EngagementReport.objects.filter(meeting=meeting).first()

    if not report and meeting.status == 'ended':
        from attendance.engagement_service import generate_engagement_report
        generate_engagement_report(meeting.id)
        report = EngagementReport.objects.filter(meeting=meeting).first()

    log_filename = f'engagement_{meeting.meeting_code}.csv'
    log_path = os.path.join(settings.MEDIA_ROOT, 'meeting_logs', log_filename)
    log_exists = os.path.exists(log_path)

    return render(request, 'attendance/engagement_report.html', {
        'meeting': meeting, 'report': report,
        'page_title': f'Engagement Report — {meeting.title}',
        'log_exists': log_exists,
        'log_url': f"{settings.MEDIA_URL}meeting_logs/{log_filename}" if log_exists else None,
    })


@login_required
def admin_face_photos(request):
    """Admin-only view of all student face registration photos."""
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access denied.")
    profiles = StudentFaceProfile.objects.filter(
        is_active=True, face_photo__isnull=False
    ).exclude(face_photo='').select_related('student', 'student__userprofile').order_by('-updated_at')
    return render(request, 'attendance/admin_face_photos.html', {
        'profiles': profiles, 'page_title': 'Student Face Photos',
    })
