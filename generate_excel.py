import openpyxl
import re
from openpyxl.styles import Font, Alignment

with open('APP_TEST_CASES.md', 'r', encoding='utf-8') as f:
    lines = f.readlines()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

row_number = 1
for line in lines:
    line = line.strip()
    if line.startswith('|') and line.endswith('|'):
        cols = [col.strip() for col in line.split('|')[1:-1]]
        if all(re.match(r'^-+$', col) for col in cols):
            continue
        cols = [col.replace('<br>', '\n') for col in cols]
        for col_index, value in enumerate(cols, 1):
            cell = ws.cell(row=row_number, column=col_index, value=value)
            # Basic wrap text if newline exists
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        row_number += 1

# Formatting
if row_number > 1:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
# Adjust column widths (don't make them too wife)
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter 
    for cell in col:
        try:
            # We don't want to size width based on multiline long text single lines
            lines_in_cell = str(cell.value).split('\n')
            for l in lines_in_cell:
                if len(l) > max_length:
                    max_length = len(l)
        except:
            pass
    adjusted_width = (max_length + 2)
    if adjusted_width > 40:
        adjusted_width = 40
    elif adjusted_width < 10:
        adjusted_width = 10
    ws.column_dimensions[column].width = adjusted_width

wb.save("APP_TEST_CASES_V2.xlsx")
print("Saved APP_TEST_CASES_V2.xlsx")
